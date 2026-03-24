"""Report generator tests for data source inventory sheet."""

from pathlib import Path

from openpyxl import load_workbook

from core.database import LocalDBManager
from report.generator import generate_report


def test_report_includes_data_source_inventory_sheet(tmp_path: Path):
    db_path = str(tmp_path / "report_inventory.db")
    mgr = LocalDBManager(db_path)
    try:
        sid = "session-report-inv"
        mgr.set_current_session_id(sid)
        mgr.create_session_record(sid)
        mgr.save_finding(
            "database",
            target_name="t",
            server_ip="127.0.0.1",
            schema_name="s",
            table_name="tb",
            column_name="col",
            data_type="TEXT",
            sensitivity_level="HIGH",
            pattern_detected="CPF",
            norm_tag="LGPD",
            ml_confidence=95,
        )
        mgr.save_data_source_inventory(
            target_name="db-primary",
            source_type="database",
            product="sqlite",
            product_version="3.45.x",
            protocol_or_api_version="sqlite",
            transport_security="local-file",
            raw_details='{"driver":"sqlite"}',
        )
        mgr.finish_session(sid)
        out = generate_report(mgr, sid, output_dir=str(tmp_path), config={})
        assert out is not None
        wb = load_workbook(out)
        assert "Data source inventory" in wb.sheetnames
        ws = wb["Data source inventory"]
        assert ws["A2"].value == "db-primary"
        assert ws["C2"].value == "sqlite"
    finally:
        mgr.dispose()
